from openpyxl import load_workbook
import logging
import pandas as pd
import requests
import json
import time

try:
    logging.basicConfig(filename='/home/vamshi/Documents/name_match/responses.log',
                        filemode='w', level=logging.DEBUG,
                        format='%(asctime)s:%(message)s')
    # cust_count = 3
    input_names = pd.read_excel('PROD_name_data.xlsx')
    cust_count = len(input_names)
    wb = load_workbook('PROD_name_data.xlsx')
    ws = wb.active

    endpoint_url = 'https://lq5rwfpsk2.execute-api.ap-south-1.amazonaws.com/r1/UATRE1_kb_name_match_validation_svc'
    headers_info = {'Content-Type': 'application/json'}

    try:
        user_set = set()
        dupe_cid = []
        for i in range(10000):

            cid = int(input_names.cid.iloc[i])

            if cid in user_set:
                dupe_cid.append(cid)
            else:
                user_set.add(cid)


        print(dupe_cid)
        print(len(dupe_cid))
            # first_name = str(input_names.first_name.iloc[i])
            # second_name = str(input_names.second_name.iloc[i])

            # req_json = {
            #     "type": "name_match_rules",
            #     "model": {
            #         "cid": str(cid),
            #         "first_name": first_name,
            #         "second_name": second_name,
            #         "source": "profile"
            #     }
            # }
            #
            # response = requests.post(endpoint_url, json=req_json, headers=headers_info)

            # cell = 'E'+str(2+i)
            # new_score = 'E' + str(2 + i)
            # response_code = 'G' + str(2 + i)
            # matched_rule = 'H' + str(2 + i)
            # ws[cell] = response.text
            # ws[new_score] = json.loads(response.text)["model"]["name_match_score"]
            # ws[response_code] = json.loads(response.text)["code"]
            # ws[matched_rule] = json.loads(response.text)["model"]["matched_rule"]
            # time.sleep(2)
            # print(i, "-->", cid, "-->", json.loads(response.text)["model"]["name_match_score"])
            # logging.debug(response.text)
            # code = json.loads(response.text).get('code', 0)
            # if code == "200":
            #     print(i, "-->", cid, "-->", json.loads(response.text)["model"]["name_match_score"])
            #
            # elif code == 0:
            #     print(i, "-->", cid, "-->", "X")
            # wb.save("PROD_name_data.xlsx")
            # time.sleep(1)

        # wb.save("PROD_name_data.xlsx")

    except Exception as e:
        print("error in the logic", e)
        # wb.save("PROD_name_data.xlsx")

except Exception as e:
    print('Error while initialing', e)

# print(response.text)