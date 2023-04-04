import numpy as np
from openpyxl import load_workbook
import pandas as pd


try:
    input_names = pd.read_excel('name_match_scores.xlsx')
    output_names = pd.read_excel('Fname_Lname_Manual_match.xlsx')
    cust_count = len(input_names)
    wb = load_workbook('Fname_Lname_Manual_match.xlsx')
    # Fname_Lname_Manual_match.xlsx
    ws = wb.active

    try:
        for i in range(30):
            name = input_names.first_name.iloc[i]
            score = output_names.Score_V1.iloc[i]
            # print(name)
            # name = 'ASDF'
            # print(output_names.loc['first_name', 'A Kumar'])
            # id = int(output_names.loc[output_names['first_name'] == name, 'S_no'].iloc[0])
            print(score, type(score), np.isnan(score))

            # old_score = 'I' + str(1 + id)
            # new_score = 'J' + str(1 + id)
            # ws[old_score] = input_names.old_score.iloc[i]
            # ws[new_score] = input_names.new_score.iloc[i]



        # wb.save("Fname_Lname_Manual_match.xlsx")



    except Exception as e:
        print("error in the logic", e)

except Exception as e:
    print('Error while initialing', e)

