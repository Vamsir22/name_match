from openpyxl import load_workbook
import pandas as pd


try:
    # input_names = pd.read_excel('name_match_scores.xlsx')
    output_names = pd.read_excel('Fname_Lname_Manual_match.xlsx')
    cust_count = len(output_names)
    wb = load_workbook('Fname_Lname_Manual_match.xlsx')
    # Fname_Lname_Manual_match.xlsx
    ws = wb.active

    try:
        for i in range():
            old_score = 'I' + str(1 + id)
            new_score = 'J' + str(1 + id)
            ws[old_score] = input_names.old_score.iloc[i]
            ws[new_score] = input_names.new_score.iloc[i]



        wb.save("Fname_Lname_Manual_match.xlsx")



    except Exception as e:
        print("error in the logic", e)

except Exception as e:
    print('Error while initialing', e)

